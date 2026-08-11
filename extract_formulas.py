import openpyxl

file_path = "/Users/macbook/Documents/alphax_crm/CUENTA MULTISPORT 2.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)

if "ALEJANDRO 2026" in wb.sheetnames:
    ws = wb["ALEJANDRO 2026"]
    
    # We are looking for row 28, 29, 30, 31 (approx, maybe offset by header lines)
    # Let's just scan all rows in column B and C where column A has text matching "cuenta"
    for r in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=r, column=1).value
        
        if isinstance(cell_a, str):
            if "carlos" in cell_a.lower() or "alejandro" in cell_a.lower() or "cuenta" in cell_a.lower():
                print(f"Row {r} | Col A: {cell_a}")
                # Print B, C, D cells (values and formulas)
                for col_idx in [2, 3, 4]:
                    cell = ws.cell(row=r, column=col_idx)
                    val = cell.value
                    if val:
                        print(f"  Col {openpyxl.utils.get_column_letter(col_idx)}: {val}")

    print("\n--- Let's also check column B for formulas just in case ---")
    for r in range(1, 40):
        cell_b = ws.cell(row=r, column=2).value
        # If it's a string starting with '=', it's a formula
        if isinstance(cell_b, str) and str(cell_b).startswith("="):
            print(f"Row {r} B: formula={cell_b}")
