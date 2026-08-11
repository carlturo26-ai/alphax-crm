import openpyxl

file_path = "/Users/macbook/Documents/alphax_crm/CUENTA MULTISPORT 2.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
ws = wb["ALEJANDRO 2026"]

rows_to_check = [17, 19, 27]
print("--- Atletas que pagaron a cuenta de Carlos (Enero) ---")
total = 0
for r in rows_to_check:
    name = ws.cell(row=r, column=1).value
    amount = ws.cell(row=r, column=2).value
    print(f"Row {r}: {name} -> ${amount}")
    if isinstance(amount, (int, float)):
        total += amount

print(f"TOTAL: ${total}")

# Now, we also need to check for "Aprendizaje" paying to Alejandro or Carlos.
# The user asked: 
# 1. "que atletas de alejandro pagaron a cuenta de carlos" -> Found for Jan?
# Let's check Feb as well.
# February is column 3 (C)
print("\n--- Febrero (Col C) Entro a cuenta de Carlos ---")
formula_c = ws.cell(row=30, column=3).value
print(f"Formula in C30 (Entro a cuenta de Carlos): {formula_c}")
if isinstance(formula_c, str) and "=" in formula_c:
    print("Need to parse this formula")
    
# What about other months?
print("\n--- Checking formulas for 'Entro a cuenta de Carlos' across months ---")
for col in range(2, 6):
    f = ws.cell(row=30, column=col).value
    print(f"Col {col} (Row 30 'Entro a cuenta de Carlos'): {f}")

print("\n--- Checking 'Aprendizaje' paying to Alejandro ---")
# "Entro neto a Alejandro" formula: row 33 
for col in range(2, 4):
    f = ws.cell(row=33, column=col).value
    print(f"Col {col} (Row 33 'Entro neto a Alejandro'): {f}")

