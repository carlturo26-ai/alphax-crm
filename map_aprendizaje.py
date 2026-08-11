import openpyxl

file_path = "/Users/macbook/Documents/alphax_crm/CUENTA MULTISPORT 2.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
ws = wb["ALEJANDRO 2026"]

print("--- Data for Aprendizaje (Row 36 to ~65) ---")
for r in range(36, 65):
    name = ws.cell(row=r, column=1).value
    val1 = ws.cell(row=r, column=2).value
    val2 = ws.cell(row=r, column=3).value
    if name:
        print(f"Row {r}: {name} | Jan(B)={val1} | Feb(C)={val2}")
        
    if isinstance(val1, str) and str(val1).startswith("="):
        print(f"  Row {r} Jan Formula: {val1}")
    if isinstance(val2, str) and str(val2).startswith("="):
        print(f"  Row {r} Feb Formula: {val2}")

